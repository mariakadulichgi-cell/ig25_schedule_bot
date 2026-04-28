import io
import os
import re
import time
import sqlite3
from datetime import date, datetime, time as dt_time, timedelta
from threading import Thread
from typing import Dict, List, Optional, Set, Tuple
from zoneinfo import ZoneInfo

import requests
from dotenv import load_dotenv
from flask import Flask
from openpyxl import load_workbook
from telegram import Update
from telegram.ext import (
    Application,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

# =========================
# Render + Flask (keep alive)
# =========================
web = Flask(__name__)


@web.get("/")
def home():
    return "ok", 200


def run_web():
    port = int(os.environ.get("PORT", "10000"))
    web.run(host="0.0.0.0", port=port)


def keep_alive():
    Thread(target=run_web, daemon=True).start()


# =========================
# Настройки
# =========================
DEFAULT_TIMEZONE = "Asia/Krasnoyarsk"
DEFAULT_GROUP = "ИГ25-01Б-ОМ"

DATE_RE = re.compile(r"\b(\d{1,2})[.\-/](\d{1,2})(?:[.\-/](\d{2,4}))?\b")
TIME_RE = re.compile(r"(\d{1,2})[.:](\d{2})\s*[–—-]\s*(\d{1,2})[.:](\d{2})")

_CACHE_ROWS: Optional[List[List[str]]] = None
_CACHE_TS: float = 0.0


# =========================
# База преподавателей
# =========================
def get_db_path() -> str:
    return os.getenv("TEACHERS_DB_PATH", "teachers.db").strip() or "teachers.db"


def parse_admin_ids() -> Set[int]:
    raw = os.getenv("ADMIN_IDS", "").strip()
    result: Set[int] = set()

    if not raw:
        return result

    for part in raw.split(","):
        part = part.strip()
        if part.isdigit():
            result.add(int(part))

    return result

def parse_teachers_from_env() -> Dict[int, str]:
    raw = os.getenv("TEACHERS", "").strip()
    result: Dict[int, str] = {}

    if not raw:
        return result

    for item in raw.split(";"):
        item = item.strip()

        if not item:
            continue

        if "=" in item:
            left, right = item.split("=", 1)
        elif ":" in item:
            left, right = item.split(":", 1)
        else:
            continue

        teacher_id = left.strip()
        fio = right.strip()

        if teacher_id.isdigit() and fio:
            result[int(teacher_id)] = fio

    return result

def is_admin(user_id: int) -> bool:
    return user_id in parse_admin_ids()


def init_teachers_db():
    conn = sqlite3.connect(get_db_path())
    cur = conn.cursor()

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS teachers (
            telegram_id INTEGER PRIMARY KEY,
            fio TEXT NOT NULL
        )
        """
    )

    conn.commit()
    conn.close()


def add_teacher_to_db(telegram_id: int, fio: str):
    conn = sqlite3.connect(get_db_path())
    cur = conn.cursor()

    cur.execute(
        """
        INSERT OR REPLACE INTO teachers (telegram_id, fio)
        VALUES (?, ?)
        """,
        (telegram_id, fio),
    )

    conn.commit()
    conn.close()


def remove_teacher_from_db(telegram_id: int) -> bool:
    conn = sqlite3.connect(get_db_path())
    cur = conn.cursor()

    cur.execute("DELETE FROM teachers WHERE telegram_id = ?", (telegram_id,))
    deleted = cur.rowcount > 0

    conn.commit()
    conn.close()

    return deleted


def get_teacher_fio(telegram_id: int) -> Optional[str]:
    teachers_from_env = parse_teachers_from_env()

    if telegram_id in teachers_from_env:
        return teachers_from_env[telegram_id]

    conn = sqlite3.connect(get_db_path())
    cur = conn.cursor()

    cur.execute("SELECT fio FROM teachers WHERE telegram_id = ?", (telegram_id,))
    row = cur.fetchone()

    conn.close()

    if row:
        return row[0]

    return None

def get_all_teachers_from_db() -> List[Tuple[int, str]]:
    result: Dict[int, str] = {}

    teachers_from_env = parse_teachers_from_env()
    result.update(teachers_from_env)

    conn = sqlite3.connect(get_db_path())
    cur = conn.cursor()

    cur.execute("SELECT telegram_id, fio FROM teachers ORDER BY fio")
    rows = cur.fetchall()

    conn.close()

    for row in rows:
        teacher_id = int(row[0])
        fio = str(row[1])

        if teacher_id not in result:
            result[teacher_id] = fio

    return sorted(result.items(), key=lambda item: item[1].lower())

# =========================
# Утилиты
# =========================
def get_tz() -> ZoneInfo:
    tz_name = os.getenv("TIMEZONE", DEFAULT_TIMEZONE).strip() or DEFAULT_TIMEZONE
    return ZoneInfo(tz_name)


def norm(value: object) -> str:
    return str(value or "").replace("\xa0", " ").strip()


def norm_group(value: object) -> str:
    text = norm(value)
    text = text.replace("—", "-").replace("–", "-")
    text = re.sub(r"\s+", " ", text)
    return text.upper()


def compact_spaces(text: str) -> str:
    text = (text or "").replace("\xa0", " ").replace("\t", " ")
    text = re.sub(r"[ ]{2,}", " ", text)
    return text.strip()


def norm_teacher_text(value: object) -> str:
    text = norm(value).lower()
    text = text.replace("ё", "е")
    text = text.replace("—", "-").replace("–", "-")
    text = re.sub(r"[^а-яa-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def parse_ddmm(text: object) -> Optional[str]:
    raw = norm(text)

    raw = raw.replace(" ", "")
    raw = re.sub(r"\.+", ".", raw)
    raw = re.sub(r"[-/]{2,}", "-", raw)
    raw = re.sub(r"[.\-/]{2,}", ".", raw)

    match = re.search(r"(\d{1,2})[.\-/](\d{1,2})(?:[.\-/](\d{2,4}))?", raw)
    if not match:
        return None

    day = int(match.group(1))
    month = int(match.group(2))

    if not (1 <= day <= 31 and 1 <= month <= 12):
        return None

    return f"{day:02d}.{month:02d}"


def normalize_time(text: object) -> Optional[str]:
    raw = norm(text)
    raw = raw.replace("—", "-").replace("–", "-")
    match = TIME_RE.search(raw)
    if not match:
        return None

    h1, m1, h2, m2 = match.groups()
    return f"{int(h1):02d}:{m1}–{int(h2):02d}:{m2}"


def to_xlsx_export_url(url: str) -> str:
    url = url.strip()
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    if not match:
        return url

    spreadsheet_id = match.group(1)
    return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"


def cell_to_text(value: object) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        if value.hour or value.minute or value.second:
            return value.strftime("%H:%M")
        return value.strftime("%d.%m.%Y")

    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")

    if isinstance(value, dt_time):
        return value.strftime("%H:%M")

    return str(value)


def worksheet_to_rows(ws) -> List[List[str]]:
    merged_values: Dict[Tuple[int, int], object] = {}

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = ws.cell(min_row, min_col).value

        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                merged_values[(row_idx, col_idx)] = top_left_value

    rows: List[List[str]] = []

    for row_idx in range(1, ws.max_row + 1):
        row: List[str] = []

        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row_idx, col_idx).value

            if value in (None, "") and (row_idx, col_idx) in merged_values:
                value = merged_values[(row_idx, col_idx)]

            row.append(cell_to_text(value))

        rows.append(row)

    return rows


def find_col_by_keywords(row_lower: List[str], keywords: List[str]) -> Optional[int]:
    for idx, cell in enumerate(row_lower):
        for keyword in keywords:
            if keyword in cell:
                return idx

    return None


def sheet_looks_like_schedule(rows: List[List[str]], group_name: str) -> bool:
    target_group = norm_group(group_name)

    for i in range(min(120, len(rows))):
        row = [norm(cell) for cell in rows[i]]
        row_lower = [cell.lower() for cell in row]

        has_date = find_col_by_keywords(row_lower, ["дата"]) is not None
        has_time = find_col_by_keywords(row_lower, ["часы", "время"]) is not None
        has_group = any(target_group == norm_group(cell) for cell in row)

        if has_date and has_time and has_group:
            return True

    return False


def fetch_sheet_rows(url: str, group_name: str, sheet_name: Optional[str] = None) -> List[List[str]]:
    export_url = to_xlsx_export_url(url)

    response = requests.get(export_url, timeout=30)
    response.raise_for_status()

    workbook = load_workbook(io.BytesIO(response.content), data_only=True)

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            raise RuntimeError(
                f"Лист '{sheet_name}' не найден. Доступные листы: {available}"
            )

        worksheet = workbook[sheet_name]
        return worksheet_to_rows(worksheet)

    for ws in workbook.worksheets:
        rows = worksheet_to_rows(ws)
        if sheet_looks_like_schedule(rows, group_name):
            return rows

    available = ", ".join(workbook.sheetnames)
    first_sheet_rows = worksheet_to_rows(workbook.worksheets[0])
    preview = "\n".join(
        " | ".join([norm(cell) for cell in first_sheet_rows[k][:12]])
        for k in range(min(10, len(first_sheet_rows)))
    )

    raise RuntimeError(
        "Не удалось автоматически найти лист с расписанием.\n"
        f"Доступные листы: {available}\n\n"
        f"Первые строки первого листа:\n{preview}"
    )


def get_rows_with_cache(url: str, group_name: str, sheet_name: Optional[str]) -> List[List[str]]:
    global _CACHE_ROWS, _CACHE_TS

    cache_seconds = int(os.getenv("CACHE_SECONDS", "60") or "60")
    now = time.time()

    if _CACHE_ROWS is not None and (now - _CACHE_TS) < cache_seconds:
        return _CACHE_ROWS

    rows = fetch_sheet_rows(url, group_name, sheet_name)
    _CACHE_ROWS = rows
    _CACHE_TS = now

    return rows


def find_header_and_group_cols(
    rows: List[List[str]],
    group_name: str,
) -> Tuple[int, int, int, List[int]]:
    target_group = norm_group(group_name)

    header_row_idx: Optional[int] = None
    date_col: Optional[int] = None
    time_col: Optional[int] = None

    for i in range(min(120, len(rows))):
        row = [norm(cell) for cell in rows[i]]
        row_lower = [cell.lower() for cell in row]

        found_date_col = find_col_by_keywords(row_lower, ["дата"])
        found_time_col = find_col_by_keywords(row_lower, ["часы", "время"])

        if found_date_col is not None and found_time_col is not None:
            header_row_idx = i
            date_col = found_date_col
            time_col = found_time_col
            break

    if header_row_idx is None or date_col is None or time_col is None:
        preview = "\n".join(
            " | ".join([norm(cell) for cell in rows[k][:12]])
            for k in range(min(10, len(rows)))
        )
        raise RuntimeError(
            "Не удалось найти заголовки 'Дата' и 'Часы/Время' в таблице.\n\n"
            f"Первые строки таблицы:\n{preview}"
        )

    group_cols: List[int] = []
    search_until = min(header_row_idx + 30, len(rows))

    for i in range(header_row_idx, search_until):
        row = rows[i]

        for j, cell in enumerate(row):
            if norm_group(cell) == target_group:
                group_cols.append(j)

    group_cols = sorted(set(group_cols))

    if not group_cols:
        for i in range(header_row_idx, search_until):
            row = rows[i]

            for j, cell in enumerate(row):
                normalized = norm_group(cell)

                if target_group and target_group in normalized:
                    group_cols.append(j)

        group_cols = sorted(set(group_cols))

    if not group_cols:
        seen_groups: Set[str] = set()

        for i in range(min(50, len(rows))):
            for cell in rows[i]:
                text = norm(cell)

                if text.upper().startswith("ИГ"):
                    seen_groups.add(text)

        hint = ", ".join(sorted(seen_groups)) if seen_groups else "ничего похожего не найдено"

        raise RuntimeError(
            f"Не нашла колонку группы '{group_name}'. В таблице нашла: {hint}"
        )

    return header_row_idx, date_col, time_col, group_cols


def should_skip_cell_text(text: str, group_name: str) -> bool:
    lowered = text.lower().strip()
    lowered = lowered.replace("ё", "е")
    lowered = re.sub(r"\s+", " ", lowered)

    if not lowered:
        return True

    if norm_group(text) == norm_group(group_name):
        return True

    trash_fragments = (
        "утверждаю",
        "семестр",
        "расписание",
        "проректор",
        "директор",
        "учебный год",
        "вид занятия",
        "аудитория",
        "вид занятия/аудитория",
        "вид занятия / аудитория",
    )

    return any(fragment in lowered for fragment in trash_fragments)

def glue_markers_to_prev(lines: List[str]) -> List[str]:
    out: List[str] = []

    for raw in lines:
        line = compact_spaces(raw)

        if not line:
            continue

        low = line.lower()

        is_marker_alone = low in {"пр", "лек", "лаб", "сем"}
        is_marker_start = bool(re.match(r"^(пр|лек|лаб|сем)\b", low))
        is_slash_room = line.startswith("/")

        if out and (is_marker_alone or is_marker_start or is_slash_room):
            out[-1] = compact_spaces(out[-1] + " " + line)
        else:
            out.append(line)

    return out


def cleanup_lines(parts: List[str], group_name: str) -> List[str]:
    text = "\n".join(parts).strip()
    raw_lines = [line.strip() for line in text.splitlines() if line.strip()]

    filtered: List[str] = []

    for line in raw_lines:
        if should_skip_cell_text(line, group_name):
            continue

        filtered.append(line)

    filtered = glue_markers_to_prev(filtered)

    uniq: List[str] = []
    seen: Set[str] = set()

    for line in filtered:
        if line not in seen:
            seen.add(line)
            uniq.append(line)

    return uniq


def time_sort_key(time_range: str) -> Tuple[int, int]:
    match = re.match(r"^(\d{2}):(\d{2})", time_range or "")

    if not match:
        return (99, 99)

    return int(match.group(1)), int(match.group(2))


def merge_items_by_time(items: List[Tuple[str, str]]) -> List[Tuple[str, str]]:
    grouped: Dict[str, List[str]] = {}
    order: List[str] = []

    for time_value, text_block in items:
        if time_value not in grouped:
            grouped[time_value] = []
            order.append(time_value)

        if text_block not in grouped[time_value]:
            grouped[time_value].append(text_block)

    ordered_times = sorted(order, key=time_sort_key)

    result: List[Tuple[str, str]] = []

    for time_value in ordered_times:
        result.append((time_value, "\n".join(grouped[time_value])))

    return result


def extract_schedule_for_date(
    rows: List[List[str]],
    group_name: str,
    target_ddmm: str,
) -> List[Tuple[str, str]]:
    header_idx, date_col, time_col, group_cols = find_header_and_group_cols(rows, group_name)

    current_date: Optional[str] = None
    current_time: Optional[str] = None
    items: List[Tuple[str, str]] = []
    seen_pairs: Set[Tuple[str, str]] = set()

    max_needed_col = max([date_col, time_col] + group_cols)

    for raw_row in rows[header_idx + 1:]:
        row = list(raw_row)

        if len(row) <= max_needed_col:
            row.extend([""] * (max_needed_col + 1 - len(row)))

        date_value = norm(row[date_col])
        time_value = norm(row[time_col])

        parsed_date = parse_ddmm(date_value)

        if parsed_date:
            current_date = parsed_date

        parsed_time = normalize_time(time_value)

        if parsed_time:
            current_time = parsed_time

        if current_date != target_ddmm:
            continue

        if not current_time:
            continue

        parts: List[str] = []

        for col in group_cols:
            value = norm(row[col])

            if not value:
                continue

            if should_skip_cell_text(value, group_name):
                continue

            parts.append(value)

        if not parts:
            continue

        lines = cleanup_lines(parts, group_name)

        if not lines:
            continue

        text_block = "\n".join(lines).strip()

        if not text_block:
            continue

        pair_key = (current_time, text_block)

        if pair_key in seen_pairs:
            continue

        seen_pairs.add(pair_key)
        items.append(pair_key)

    return merge_items_by_time(items)


def format_schedule(group_name: str, ddmm: str, items: List[Tuple[str, str]]) -> str:
    months = {
        "01": "января",
        "02": "февраля",
        "03": "марта",
        "04": "апреля",
        "05": "мая",
        "06": "июня",
        "07": "июля",
        "08": "августа",
        "09": "сентября",
        "10": "октября",
        "11": "ноября",
        "12": "декабря",
    }

    day, month = ddmm.split(".")
    pretty_date = f"{int(day)} {months.get(month, month)}"

    if not items:
        return (
            f"📅 Расписание на {pretty_date}\n"
            f"━━━━━━━━━━━━━━\n\n"
            f"На этот день пар нет 💙"
        )

    out_lines: List[str] = [
        f"📅 Расписание на {pretty_date}",
        "━━━━━━━━━━━━━━",
        ""
    ]

    for time_value, text_block in items:
        lines = [line.strip() for line in text_block.splitlines() if line.strip()]
        lines = glue_markers_to_prev(lines)

        if not lines:
            continue

        subject = lines[0]
        details = lines[1:]

        out_lines.append(f"⏰ {time_value}")
        out_lines.append(f"📌 {subject}")

        if details:
            out_lines.append(f"👥 {' | '.join(details)}")

        out_lines.append("")

    while out_lines and out_lines[-1] == "":
        out_lines.pop()

    return "\n".join(out_lines)

# =========================
# Расписание преподавателя
# =========================
def looks_like_group_name(value: object) -> bool:
    text = norm_group(value)

    if not text:
        return False

    if text.startswith("ИГ") and any(ch.isdigit() for ch in text):
        return True

    if re.search(r"\b[А-ЯЁA-Z]{1,6}\d{2}-\d{2}", text):
        return True

    return False


def find_header_and_all_group_cols(
    rows: List[List[str]],
) -> Tuple[int, int, int, Dict[str, List[int]]]:
    header_row_idx: Optional[int] = None
    date_col: Optional[int] = None
    time_col: Optional[int] = None

    for i in range(min(120, len(rows))):
        row = [norm(cell) for cell in rows[i]]
        row_lower = [cell.lower() for cell in row]

        found_date_col = find_col_by_keywords(row_lower, ["дата"])
        found_time_col = find_col_by_keywords(row_lower, ["часы", "время"])

        if found_date_col is not None and found_time_col is not None:
            header_row_idx = i
            date_col = found_date_col
            time_col = found_time_col
            break

    if header_row_idx is None or date_col is None or time_col is None:
        raise RuntimeError("Не удалось найти заголовки 'Дата' и 'Часы/Время' в таблице.")

    group_to_cols: Dict[str, List[int]] = {}
    search_until = min(header_row_idx + 30, len(rows))

    for i in range(header_row_idx, search_until):
        row = rows[i]

        for j, cell in enumerate(row):
            if j in {date_col, time_col}:
                continue

            if looks_like_group_name(cell):
                group_name = compact_spaces(norm(cell))
                group_to_cols.setdefault(group_name, [])

                if j not in group_to_cols[group_name]:
                    group_to_cols[group_name].append(j)

    if not group_to_cols:
        raise RuntimeError(
            "Не удалось найти группы в таблице. "
            "Проверь, что названия групп находятся рядом с заголовками расписания."
        )

    for group_name in list(group_to_cols.keys()):
        group_to_cols[group_name] = sorted(set(group_to_cols[group_name]))

    return header_row_idx, date_col, time_col, group_to_cols


def extract_teacher_schedule_for_date(
    rows: List[List[str]],
    teacher_fio: str,
    target_ddmm: str,
) -> List[Tuple[str, str, str]]:
    header_idx, date_col, time_col, group_to_cols = find_header_and_all_group_cols(rows)

    teacher_norm = norm_teacher_text(teacher_fio)

    current_date: Optional[str] = None
    current_time: Optional[str] = None
    items: List[Tuple[str, str, str]] = []
    seen: Set[Tuple[str, str, str]] = set()

    all_cols: List[int] = [date_col, time_col]

    for cols in group_to_cols.values():
        all_cols.extend(cols)

    max_needed_col = max(all_cols)

    for raw_row in rows[header_idx + 1:]:
        row = list(raw_row)

        if len(row) <= max_needed_col:
            row.extend([""] * (max_needed_col + 1 - len(row)))

        date_value = norm(row[date_col])
        time_value = norm(row[time_col])

        parsed_date = parse_ddmm(date_value)

        if parsed_date:
            current_date = parsed_date

        parsed_time = normalize_time(time_value)

        if parsed_time:
            current_time = parsed_time

        if current_date != target_ddmm:
            continue

        if not current_time:
            continue

        for group_name, group_cols in group_to_cols.items():
            parts: List[str] = []

            for col in group_cols:
                value = norm(row[col])

                if not value:
                    continue

                if should_skip_cell_text(value, group_name):
                    continue

                parts.append(value)

            if not parts:
                continue

            lines = cleanup_lines(parts, group_name)

            if not lines:
                continue

            text_block = "\n".join(lines).strip()

            if not text_block:
                continue

            if teacher_norm not in norm_teacher_text(text_block):
                continue

            key = (current_time, group_name, text_block)

            if key in seen:
                continue

            seen.add(key)
            items.append(key)

    items.sort(key=lambda item: time_sort_key(item[0]))

    return items


def format_teacher_schedule(
    teacher_fio: str,
    ddmm: str,
    items: List[Tuple[str, str, str]],
) -> str:
    months = {
        "01": "января",
        "02": "февраля",
        "03": "марта",
        "04": "апреля",
        "05": "мая",
        "06": "июня",
        "07": "июля",
        "08": "августа",
        "09": "сентября",
        "10": "октября",
        "11": "ноября",
        "12": "декабря",
    }

    day, month = ddmm.split(".")
    pretty_date = f"{int(day)} {months.get(month, month)}"

    if not items:
        return (
            f"📅 Расписание преподавателя на {pretty_date}\n"
            f"👤 {teacher_fio}\n"
            f"━━━━━━━━━━━━━━\n\n"
            f"На этот день пар нет 💙"
        )

    out_lines: List[str] = [
        f"📅 Расписание преподавателя на {pretty_date}",
        f"👤 {teacher_fio}",
        "━━━━━━━━━━━━━━",
        ""
    ]

    for time_value, group_name, text_block in items:
        lines = [line.strip() for line in text_block.splitlines() if line.strip()]
        lines = glue_markers_to_prev(lines)

        if not lines:
            continue

        subject = lines[0]
        details = lines[1:]

        out_lines.append(f"⏰ {time_value}")
        out_lines.append(f"👥 {group_name}")
        out_lines.append(f"📌 {subject}")

        if details:
            out_lines.append(f"📝 {' | '.join(details)}")

        out_lines.append("")

    while out_lines and out_lines[-1] == "":
        out_lines.pop()

    return "\n".join(out_lines)

def parse_teacher_date_from_args(args: List[str]) -> Optional[str]:
    raw = " ".join(args).strip().lower()

    if not raw:
        return datetime.now(get_tz()).strftime("%d.%m")

    if raw in {"сегодня", "today"}:
        return datetime.now(get_tz()).strftime("%d.%m")

    if raw in {"завтра", "tomorrow"}:
        return (datetime.now(get_tz()) + timedelta(days=1)).strftime("%d.%m")

    return parse_ddmm(raw)


# =========================
# Сообщения
# =========================
def split_message(text: str, limit: int = 3800) -> List[str]:
    if len(text) <= limit:
        return [text]

    chunks: List[str] = []
    current = ""

    for line in text.splitlines(True):
        if len(current) + len(line) <= limit:
            current += line
            continue

        if current.strip():
            chunks.append(current.rstrip())
            current = ""

        if len(line) <= limit:
            current = line
            continue

        rest = line

        while len(rest) > limit:
            chunks.append(rest[:limit])
            rest = rest[limit:]

        current = rest

    if current.strip():
        chunks.append(current.rstrip())

    return chunks


async def reply_long(update: Update, text: str) -> None:
    if not update.message:
        return

    for chunk in split_message(text):
        await update.message.reply_text(chunk)


# =========================
# Telegram handlers
# =========================
async def send_schedule(update: Update, ddmm: str):
    sheet_url = os.getenv("SHEET_URL", "").strip()
    sheet_name = os.getenv("SHEET_NAME", "").strip() or None
    group_name = os.getenv("GROUP_NAME", DEFAULT_GROUP).strip() or DEFAULT_GROUP

    if not sheet_url:
        await reply_long(update, "Не задана переменная SHEET_URL.")
        return

    try:
        rows = get_rows_with_cache(sheet_url, group_name, sheet_name)
        items = extract_schedule_for_date(rows, group_name, ddmm)
        message = format_schedule(group_name, ddmm, items)
        await reply_long(update, message)

    except Exception as exc:
        await reply_long(update, f"Ошибка чтения расписания: {exc}")


async def send_teacher_schedule(update: Update, teacher_fio: str, ddmm: str):
    sheet_url = os.getenv("SHEET_URL", "").strip()
    sheet_name = os.getenv("SHEET_NAME", "").strip() or None
    group_name = os.getenv("GROUP_NAME", DEFAULT_GROUP).strip() or DEFAULT_GROUP

    if not sheet_url:
        await reply_long(update, "Не задана переменная SHEET_URL.")
        return

    try:
        rows = get_rows_with_cache(sheet_url, group_name, sheet_name)
        items = extract_teacher_schedule_for_date(rows, teacher_fio, ddmm)
        message = format_teacher_schedule(teacher_fio, ddmm, items)
        await reply_long(update, message)

    except Exception as exc:
        await reply_long(update, f"Ошибка чтения расписания преподавателя: {exc}")


async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (
        "Привет! Я бот с расписанием 💙\n\n"
        "Команды для студентов:\n"
        "/today — расписание на сегодня\n"
        "/tomorrow — расписание на завтра\n"
        "/day 30.01 — расписание на дату\n\n"
        "Можно просто написать: 30.01\n"
        "Или: день 30.01\n\n"
        "Команды для преподавателей:\n"
        "/id — узнать свой Telegram ID\n"
        "/prepod — расписание преподавателя на сегодня\n"
        "/prepod завтра — расписание преподавателя на завтра\n"
        "/prepod 30.01 — расписание преподавателя на дату"
    )

    await reply_long(update, text)


async def cmd_today(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ddmm = datetime.now(get_tz()).strftime("%d.%m")
    await send_schedule(update, ddmm)


async def cmd_tomorrow(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ddmm = (datetime.now(get_tz()) + timedelta(days=1)).strftime("%d.%m")
    await send_schedule(update, ddmm)


async def cmd_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = " ".join(context.args).strip()

    if not raw:
        await reply_long(update, "Напиши так: /day 30.01")
        return

    ddmm = parse_ddmm(raw)

    if not ddmm:
        await reply_long(update, "Формат даты: /day 30.01")
        return

    await send_schedule(update, ddmm)


async def cmd_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.effective_user:
        return

    user = update.effective_user
    username = f"@{user.username}" if user.username else "username не указан"

    text = (
        f"Ваш Telegram ID: {user.id}\n"
        f"Ваш username: {username}\n\n"
        f"Отправьте этот ID администратору, чтобы получить доступ к расписанию преподавателя."
    )

    await reply_long(update, text)


async def cmd_add_teacher(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.effective_user:
        return

    user_id = update.effective_user.id

    if not is_admin(user_id):
        await reply_long(update, "У вас нет доступа к этой команде.")
        return

    if len(context.args) < 2:
        await reply_long(
            update,
            "Использование команды:\n"
            "/add_teacher 123456789 Иванова Мария Петровна\n\n"
            "Важно: ФИО нужно писать так же, как оно написано в таблице.",
        )
        return

    try:
        teacher_id = int(context.args[0])
    except ValueError:
        await reply_long(update, "Telegram ID должен быть числом.")
        return

    fio = " ".join(context.args[1:]).strip()

    if not fio:
        await reply_long(update, "Напиши ФИО преподавателя после Telegram ID.")
        return

    add_teacher_to_db(teacher_id, fio)

    await reply_long(
        update,
        f"Преподаватель добавлен:\n"
        f"{fio}\n"
        f"Telegram ID: {teacher_id}",
    )


async def cmd_remove_teacher(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.effective_user:
        return

    user_id = update.effective_user.id

    if not is_admin(user_id):
        await reply_long(update, "У вас нет доступа к этой команде.")
        return

    if len(context.args) < 1:
        await reply_long(update, "Использование команды:\n/remove_teacher 123456789")
        return

    try:
        teacher_id = int(context.args[0])
    except ValueError:
        await reply_long(update, "Telegram ID должен быть числом.")
        return

    deleted = remove_teacher_from_db(teacher_id)

    if deleted:
        await reply_long(update, f"Преподаватель с ID {teacher_id} удалён.")
    else:
        await reply_long(update, f"Преподаватель с ID {teacher_id} не найден в базе.")


async def cmd_list_teachers(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.effective_user:
        return

    user_id = update.effective_user.id

    if not is_admin(user_id):
        await reply_long(update, "У вас нет доступа к этой команде.")
        return

    teachers = get_all_teachers_from_db()

    if not teachers:
        await reply_long(update, "В базе пока нет преподавателей.")
        return

    lines = ["Преподаватели в базе:"]

    for teacher_id, fio in teachers:
        lines.append(f"• {fio} — {teacher_id}")

    await reply_long(update, "\n".join(lines))


async def get_teacher_fio_or_reply(update: Update) -> Optional[str]:
    if not update.effective_user:
        return None

    user_id = update.effective_user.id
    teacher_fio = get_teacher_fio(user_id)

    if not teacher_fio:
        await reply_long(
            update,
            "У вас нет доступа к расписанию преподавателя.\n\n"
            "Сначала отправьте администратору команду /id.",
        )
        return None

    return teacher_fio


async def cmd_prepod_today(update: Update, context: ContextTypes.DEFAULT_TYPE):
    teacher_fio = await get_teacher_fio_or_reply(update)

    if not teacher_fio:
        return

    ddmm = datetime.now(get_tz()).strftime("%d.%m")
    await send_teacher_schedule(update, teacher_fio, ddmm)


async def cmd_prepod_tomorrow(update: Update, context: ContextTypes.DEFAULT_TYPE):
    teacher_fio = await get_teacher_fio_or_reply(update)

    if not teacher_fio:
        return

    ddmm = (datetime.now(get_tz()) + timedelta(days=1)).strftime("%d.%m")
    await send_teacher_schedule(update, teacher_fio, ddmm)


async def cmd_prepod_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    teacher_fio = await get_teacher_fio_or_reply(update)

    if not teacher_fio:
        return

    raw = " ".join(context.args).strip()

    if not raw:
        await reply_long(update, "Напишите так: /prepod_day 30.01")
        return

    ddmm = parse_ddmm(raw)

    if not ddmm:
        await reply_long(update, "Формат даты: /prepod_day 30.01")
        return

    await send_teacher_schedule(update, teacher_fio, ddmm)
    if not ddmm:
        await reply_long(
            update,
            "Формат команды:\n"
            "/prepod\n"
            "/prepod завтра\n"
            "/prepod 30.01",
        )
        return

    await send_teacher_schedule(update, teacher_fio, ddmm)


async def text_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message:
        return

    text = (update.message.text or "").strip().lower()

    match = re.match(
        r"^(?:день\s+)?(\d{1,2}[.\-/]\d{1,2}(?:[.\-/]\d{2,4})?)$",
        text,
    )

    if not match:
        return

    ddmm = parse_ddmm(match.group(1))

    if not ddmm:
        return

    await send_schedule(update, ddmm)


# =========================
# Main
# =========================
def main():
    load_dotenv()

    token = os.getenv("BOT_TOKEN", "").strip()

    if not token:
        raise RuntimeError("Нет BOT_TOKEN в .env")

    init_teachers_db()
    keep_alive()

    app = Application.builder().token(token).build()

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("today", cmd_today))
    app.add_handler(CommandHandler("tomorrow", cmd_tomorrow))
    app.add_handler(CommandHandler("day", cmd_day))

    app.add_handler(CommandHandler("id", cmd_id))
    app.add_handler(CommandHandler("add_teacher", cmd_add_teacher))
    app.add_handler(CommandHandler("remove_teacher", cmd_remove_teacher))
    app.add_handler(CommandHandler("list_teachers", cmd_list_teachers))
    app.add_handler(CommandHandler("prepod_today", cmd_prepod_today))
    app.add_handler(CommandHandler("prepod_tomorrow", cmd_prepod_tomorrow))
    app.add_handler(CommandHandler("prepod_day", cmd_prepod_day))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_day))

    print("Bot started / polling")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
